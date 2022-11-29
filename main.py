import pandas as pd
import openpyxl
import prod_part
from options import path_to_project
from datetime import datetime
now = datetime.now()

file = openpyxl.open(rf"{path_to_project}\settings_file.xlsx", read_only=True)

sheet = file.sheetnames
print(sheet)
print(len(sheet))
count = 1
file_names=[]
for i in sheet:
    print(f"ДАННЫЕ ИЗ ЛИСТА: {i}")
    sheet = file[i]
    # print("Продукт: ")
    name_prod = sheet[2][0].value

    # print("Толщины: ")
    thinkness_prod = []
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=3):
        thinkness_prod.append(row[0].value)
    thinkness_prod = [i for i in thinkness_prod if i is not None]
    # print(thinkness_prod)

    # print("Города: ")
    cities = []
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=4):
        cities.append(row[0].value)
    cities = [i for i in cities if i is not None]

    print(f"Продукт: {name_prod}")
    print(f"Толщины: {thinkness_prod}")
    print(f'Города: {cities}')
    print()

    file_name=prod_part.main_prod(name_prod, thinkness_prod, cities, count,i)
    file_names.append(file_name)
    count += 1

print()
print("Ваши файлы: (лежат в папке Excel)")
for file_name in file_names:
    print(file_name)

